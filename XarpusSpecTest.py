import sys
import os
from datetime import datetime

# Add the app directory to the Python path FIRST
app_dir = os.path.join(os.path.dirname(__file__), 'app')
sys.path.insert(0, app_dir)

# Pre-load the correct modules and create aliases in sys.modules
# This ensures all submodules use the same class instances
import NPC
import Weapon
import Stats
import Player

# Create module aliases so Weapons.Weapon, Monsters.NPC, etc. all resolve to the same modules
sys.modules['Monsters.NPC'] = sys.modules['NPC']
sys.modules['Weapons.Weapon'] = sys.modules['Weapon']
sys.modules['Weapons.Stats'] = sys.modules['Stats']
sys.modules['Weapons.NPC'] = sys.modules['NPC']
sys.modules['Loadouts.Stats'] = sys.modules['Stats']
sys.modules['Loadouts.Weapon'] = sys.modules['Weapon']
sys.modules['Loadouts.Player'] = sys.modules['Player']

from Monsters.Xarpus import Xarpus
from Monsters.Nylo_boss import P3Verzik as Nylo_boss
from Monsters.Maiden import Maiden
from Monsters.Sotetseg import Sotetseg

from Player import Player
from Weapons.Scythe import Scythe
from Weapons.NoxHally import NoxHally
from Weapons.Fists import Fists
from Weapons.Bgs import Bgs
from Weapons.ElderMaul import ElderMaul 
from Weapons.DragonClaws import DragonClaws
from Weapons.CrystalHalberd import CrystalHalberd



from Loadouts.OathTorvaRancour import player as oath_torva_rancour
from Loadouts.OathTorvaSalve import player as oath_torva_salve


number_of_players = 5

scythe = Scythe()
nox = NoxHally()
fists = Fists()
bgs = Bgs()
dragon_claws = DragonClaws()
eldermaul = ElderMaul()
crystal_halberd = CrystalHalberd()  # Using actual CrystalHalberd class

scythe_attack_speed = 5  # Attack speed in ticks (5 ticks = 3.0 seconds)

player_names = ["1sfrz", "2rdps", "3mdps", "4nrfz", "5mdps"]

# Weapon mapping for user input
weapon_options = {
    '1': ('Elder Maul', eldermaul),
    '2': ('BGS', bgs),
    '3': ('Dragon Claws', dragon_claws),
    '4': ('Crystal Halberd', crystal_halberd),
    '5': ('Scythe (regular hit)', scythe),
}

# Boss mapping for user input
boss_options = {
    '1': ('Xarpus', Xarpus),
    '2': ('Nylo Boss', Nylo_boss),
    '3': ('Maiden', Maiden),
    '4': ('Sotetseg', Sotetseg),
}


def get_user_input():
    """
    Get user input for boss, number of players, rounds, and weapon specs.
    
    Returns:
        Tuple of (boss_class, num_players, player_specs)
    """
    print("\n" + "="*50)
    print("TOB BOSS SPEC SIMULATION SETUP")
    print("="*50)
    
    # Get boss selection
    print("\n" + "-"*50)
    print("BOSS OPTIONS:")
    for key, (name, _) in boss_options.items():
        print(f"  {key}. {name}")
    print("-"*50)
    
    while True:
        boss_choice = input("\nSelect boss (1-4): ")
        if boss_choice in boss_options:
            boss_name, boss_class = boss_options[boss_choice]
            print(f"  → Selected: {boss_name}")
            break
        else:
            print("Invalid choice. Please enter 1-4.")
    
    # Get number of players
    while True:
        try:
            num_players = int(input("\nEnter number of players (3-5): "))
            if 3 <= num_players <= 5:
                break
            else:
                print("Please enter a number between 3 and 5.")
        except ValueError:
            print("Invalid input. Please enter a number.")
    
    # Special handling for Sotetseg - fixed 3 waves of 2 rounds each
    if boss_name == 'Sotetseg':
        print("\nSotetseg has 3 WAVES with 2 rounds of specs each")
        num_waves = 3
        rounds_per_wave = 2
        total_rounds = num_waves * rounds_per_wave
    else:
        # All other bosses have 2 rounds of specs
        print(f"\n{boss_name} has 2 rounds of specs")
        total_rounds = 2
    
    # Display weapon options with numbers
    print("\n" + "-"*50)
    print("WEAPON OPTIONS:")
    for key, (name, _) in weapon_options.items():
        print(f"  [{key}] {name}")
    print("-"*50)
    
    # Get weapon choices for each round and each player
    player_specs = []
    
    if boss_name == 'Sotetseg':
        # Sotetseg: 3 waves, 2 rounds each
        for wave_num in range(num_waves):
            print(f"\n{'*'*50}")
            print(f"WAVE {wave_num + 1} of {num_waves}")
            print('*'*50)
            for round_num in range(rounds_per_wave):
                print(f"\n--- Wave {wave_num + 1}, Round {round_num + 1} ---")
                round_specs = []
                for player_num in range(num_players):
                    while True:
                        choice = input(f"Player {player_names[player_num]} weapon: ")
                        if choice in weapon_options:
                            weapon_name, weapon = weapon_options[choice]
                            round_specs.append(weapon)
                            print(f"  → {player_names[player_num]} will use {weapon_name}")
                            break
                        else:
                            print("Invalid choice. Please enter 1-5.")
                player_specs.append(round_specs)
    else:
        # Regular bosses
        for round_num in range(total_rounds):
            print(f"\n--- Round {round_num + 1} ---")
            round_specs = []
            for player_num in range(num_players):
                while True:
                    choice = input(f"Player {player_names[player_num]} weapon: ")
                    if choice in weapon_options:
                        weapon_name, weapon = weapon_options[choice]
                        round_specs.append(weapon)
                        print(f"  → {player_names[player_num]} will use {weapon_name}")
                        break
                    else:
                        print("Invalid choice. Please enter 1-5.")
            player_specs.append(round_specs)
    
    return boss_class, num_players, player_specs


def run_spec_simulation(boss_class, num_players, player_specs, num_simulations=1, verbose=True):
    """
    Run a spec simulation with multiple rounds, repeated multiple times.
    
    Args:
        boss_class: The boss class to instantiate (Xarpus, Bloat, etc.)
        num_players: Number of players in the fight (affects boss stats)
        player_specs: List of rounds, where each round is a list of weapons for each player
                     Example: [[eldermaul, eldermaul, bgs, bgs, dragon_claws],  # Round 1
                               [bgs, bgs, bgs, bgs, bgs]]                         # Round 2
                     Or single round: [eldermaul, eldermaul, bgs, bgs, dragon_claws]
        num_simulations: How many times to run the simulation
        verbose: If True, print details of each simulation
    
    Returns:
        Dictionary containing statistics about damage, defense reduction, etc.
    """
    # Check if player_specs is a single round or multiple rounds
    if isinstance(player_specs[0], list):
        # Multiple rounds
        spec_rounds = player_specs
    else:
        # Single round - wrap it in a list
        spec_rounds = [player_specs]
    
    results = {
        'total_damage': [],
        'final_defense': [],
        'defense_after_each_round': [[] for _ in spec_rounds],
        'damage_per_round': [[] for _ in spec_rounds],
        'damage_per_player': {name: [] for name in player_names[:num_players]},
        'spec_hits_per_player': {name: [] for name in player_names[:num_players]},
        'def_100_or_less_round_1': 0,  # Count how many times round 1 reaches 100 or less
        'def_15_or_less_round_2': 0    # Count how many times round 2 reaches 15 or less
    }
    
    for sim in range(num_simulations):
        if verbose and num_simulations > 1:
            print(f"\n{'='*50}")
            print(f"Simulation {sim + 1}/{num_simulations}")
            print('='*50)
        
        # Create fresh players for this simulation
        players = []
        for i in range(num_players):
            player = Player(stats=oath_torva_rancour.stats.get_stats())
            player.current_special_attack = 100  # Reset spec to 100 for each simulation
            players.append((player_names[i], player))
        
        # Create fresh boss
        boss = boss_class(num_players)
        
        starting_def = int(boss.stats.def_level)
        if verbose:
            print(f"Starting defense level: {starting_def}")
        
        # Track damage per player this simulation
        sim_damage = {name: 0 for name, _ in players}
        sim_total_damage = 0
        
        # Run each round of specs
        for round_num, round_specs in enumerate(spec_rounds):
            round_damage = 0
            if verbose:
                print(f"\n--- Round {round_num + 1} specs ---")
            
            # Each player does their spec for this round
            for i, (name, player) in enumerate(players):
                player.equip_weapon(round_specs[i])
                # Check if weapon is scythe (regular hit, not special)
                if round_specs[i].__class__.__name__ == 'Scythe':
                    hit = player.do_attack(boss, special_attack=False)
                else:
                    hit = player.do_attack(boss, special_attack=True)
                boss.reduce_hp(hit)
                sim_damage[name] += hit
                round_damage += hit
                if verbose:
                    weapon_name = player.weapon.__class__.__name__
                    attack_type = "hit" if round_specs[i].__class__.__name__ == 'Scythe' else "spec"
                    print(f"{name} {weapon_name} {attack_type}: {hit}")
            
            defense_after_round = int(boss.stats.def_level)
            if verbose:
                print(f"Defense after round {round_num + 1}: {defense_after_round}")
            
            # Store round results
            results['defense_after_each_round'][round_num].append(defense_after_round)
            results['damage_per_round'][round_num].append(round_damage)
            sim_total_damage += round_damage
            
            # Track specific defense thresholds
            if round_num == 0 and defense_after_round <= 100:
                results['def_100_or_less_round_1'] += 1
            if round_num == 1 and defense_after_round <= 15:
                results['def_15_or_less_round_2'] += 1
        
        final_def = int(boss.stats.def_level)
        if verbose:
            print(f"\nFinal defense level: {final_def}")
            print(f"Total damage dealt: {sim_total_damage}")
        
        # Store results
        results['total_damage'].append(sim_total_damage)
        results['final_defense'].append(final_def)
        for name in player_names[:num_players]:
            results['damage_per_player'][name].append(sim_damage[name])
    
    # Calculate and print summary statistics if multiple simulations
    if num_simulations > 1:
        print(f"\n{'='*50}")
        print(f"SUMMARY STATISTICS ({num_simulations} simulations)")
        print('='*50)
        
        for round_num in range(len(spec_rounds)):
            avg_def = sum(results['defense_after_each_round'][round_num]) / num_simulations
            min_def = min(results['defense_after_each_round'][round_num])
            max_def = max(results['defense_after_each_round'][round_num])
            avg_dmg = sum(results['damage_per_round'][round_num]) / num_simulations
            print(f"\nRound {round_num + 1}:")
            print(f"  Average defense after: {avg_def:.2f}")
            print(f"  Min defense: {min_def}")
            print(f"  Max defense: {max_def}")
            print(f"  Average damage dealt: {avg_dmg:.2f}")
            
            # Show threshold percentages
            if round_num == 0:
                percent_100_or_less = (results['def_100_or_less_round_1'] / num_simulations) * 100
                print(f"  % of time defense ≤ 100: {percent_100_or_less:.1f}%")
            if round_num == 1:
                percent_15_or_less = (results['def_15_or_less_round_2'] / num_simulations) * 100
                print(f"  % of time defense ≤ 15: {percent_15_or_less:.1f}%")
        
        print(f"\nAverage total damage (all rounds): {sum(results['total_damage']) / num_simulations:.2f}")
        print(f"Average final defense: {sum(results['final_defense']) / num_simulations:.2f}")
        
        print(f"\nDamage by player (total average):")
        total_team_damage = 0
        for name in player_names[:num_players]:
            avg_dmg = sum(results['damage_per_player'][name]) / num_simulations
            total_team_damage += avg_dmg
            print(f"  {name}: {avg_dmg:.2f}")
        
        print(f"\nTotal team damage (average): {total_team_damage:.2f}")
    
    return results


def export_to_excel(results, boss_name, num_players, player_specs, num_simulations):
    """
    Export simulation results to an Excel spreadsheet.
    
    Args:
        results: Dictionary containing simulation results
        boss_name: Name of the boss
        num_players: Number of players
        player_specs: List of weapon specs per round
        num_simulations: Number of simulations run
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("\n⚠️  openpyxl library not found. Installing...")
        import subprocess
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Simulation Results"
    
    # Styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    subheader_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    subheader_font = Font(bold=True)
    
    row = 1
    
    # Title
    ws[f'A{row}'] = f"TOB {boss_name} Spec Simulation Results"
    ws[f'A{row}'].font = Font(size=14, bold=True)
    row += 1
    
    # Metadata
    ws[f'A{row}'] = f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    row += 1
    ws[f'A{row}'] = f"Number of Players: {num_players}"
    row += 1
    ws[f'A{row}'] = f"Number of Simulations: {num_simulations}"
    row += 2
    
    # Spec rounds configuration
    ws[f'A{row}'] = "Spec Configuration"
    ws[f'A{row}'].font = subheader_font
    ws[f'A{row}'].fill = subheader_fill
    row += 1
    
    for round_num, round_specs in enumerate(player_specs):
        ws[f'A{row}'] = f"Round {round_num + 1}:"
        spec_names = [spec.__class__.__name__ for spec in round_specs]
        ws[f'B{row}'] = ", ".join(spec_names)
        row += 1
    row += 1
    
    # Summary Statistics Header
    ws[f'A{row}'] = "Summary Statistics"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws.merge_cells(f'A{row}:C{row}')
    row += 1
    
    # Column headers
    ws[f'A{row}'] = "Metric"
    ws[f'B{row}'] = "Average"
    ws[f'C{row}'] = "Min"
    ws[f'D{row}'] = "Max"
    for col in ['A', 'B', 'C', 'D']:
        ws[f'{col}{row}'].font = subheader_font
        ws[f'{col}{row}'].fill = subheader_fill
    row += 1
    
    # Total damage stats
    ws[f'A{row}'] = "Total Damage (all rounds)"
    ws[f'B{row}'] = sum(results['total_damage']) / num_simulations
    ws[f'C{row}'] = min(results['total_damage'])
    ws[f'D{row}'] = max(results['total_damage'])
    row += 1
    
    # Final defense stats
    ws[f'A{row}'] = "Final Defense Level"
    ws[f'B{row}'] = sum(results['final_defense']) / num_simulations
    ws[f'C{row}'] = min(results['final_defense'])
    ws[f'D{row}'] = max(results['final_defense'])
    row += 2
    
    # Round-by-round stats
    ws[f'A{row}'] = "Round-by-Round Results"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws.merge_cells(f'A{row}:D{row}')
    row += 1
    
    for round_num in range(len(player_specs)):
        ws[f'A{row}'] = f"Round {round_num + 1}"
        ws[f'A{row}'].font = subheader_font
        row += 1
        
        ws[f'A{row}'] = "  Avg Defense After"
        ws[f'B{row}'] = sum(results['defense_after_each_round'][round_num]) / num_simulations
        row += 1
        
        ws[f'A{row}'] = "  Avg Damage Dealt"
        ws[f'B{row}'] = sum(results['damage_per_round'][round_num]) / num_simulations
        row += 1
        
        # Defense thresholds
        if round_num == 0:
            percent_100 = (results['def_100_or_less_round_1'] / num_simulations) * 100
            ws[f'A{row}'] = "  % Defense ≤ 100"
            ws[f'B{row}'] = f"{percent_100:.1f}%"
            row += 1
        if round_num == 1:
            percent_15 = (results['def_15_or_less_round_2'] / num_simulations) * 100
            ws[f'A{row}'] = "  % Defense ≤ 15"
            ws[f'B{row}'] = f"{percent_15:.1f}%"
            row += 1
    
    row += 1
    
    # Player damage breakdown
    ws[f'A{row}'] = "Player Damage Breakdown"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws.merge_cells(f'A{row}:C{row}')
    row += 1
    
    ws[f'A{row}'] = "Player"
    ws[f'B{row}'] = "Average Damage"
    ws[f'C{row}'] = "% of Total"
    for col in ['A', 'B', 'C']:
        ws[f'{col}{row}'].font = subheader_font
        ws[f'{col}{row}'].fill = subheader_fill
    row += 1
    
    total_team_damage = sum([sum(results['damage_per_player'][name]) / num_simulations 
                             for name in player_names[:num_players]])
    
    for name in player_names[:num_players]:
        avg_dmg = sum(results['damage_per_player'][name]) / num_simulations
        percent = (avg_dmg / total_team_damage * 100) if total_team_damage > 0 else 0
        ws[f'A{row}'] = name
        ws[f'B{row}'] = avg_dmg
        ws[f'C{row}'] = f"{percent:.1f}%"
        row += 1
    
    ws[f'A{row}'] = "Total Team Damage"
    ws[f'A{row}'].font = subheader_font
    ws[f'B{row}'] = total_team_damage
    ws[f'B{row}'].font = subheader_font
    row += 2
    
    # Raw data sheet
    ws2 = wb.create_sheet("Raw Data")
    ws2[f'A1'] = "Simulation #"
    ws2[f'B1'] = "Total Damage"
    ws2[f'C1'] = "Final Defense"
    for col in ['A', 'B', 'C']:
        ws2[f'{col}1'].font = subheader_font
        ws2[f'{col}1'].fill = subheader_fill
    
    for i in range(num_simulations):
        ws2[f'A{i+2}'] = i + 1
        ws2[f'B{i+2}'] = results['total_damage'][i]
        ws2[f'C{i+2}'] = results['final_defense'][i]
    
    # Adjust column widths
    for sheet in [ws, ws2]:
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    # Generate filename
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"TOB_{boss_name}_Sim_{num_players}p_{num_simulations}runs_{timestamp}.xlsx"
    
    # Save workbook
    wb.save(filename)
    print(f"\n✅ Results exported to: {filename}")
    return filename


print()
print("Running TOB Boss Spec Simulation")
print("="*50)

# Get user input
boss_class, num_players, player_specs = get_user_input()

# Ask for number of simulations
while True:
    try:
        num_sims = int(input("\nHow many simulations to run? (1-10000): "))
        if 1 <= num_sims <= 10000:
            break
        else:
            print("Please enter a number between 1 and 10000.")
    except ValueError:
        print("Invalid input. Please enter a number.")

# Determine verbosity based on number of simulations
verbose = num_sims <= 10  # Only show details if running 10 or fewer

print(f"\nRunning {num_sims} simulation(s) with {num_players} players...")

# Run the simulation
results = run_spec_simulation(
    boss_class=boss_class,
    num_players=num_players,
    player_specs=player_specs,
    num_simulations=num_sims,
    verbose=verbose
)

# # Ask if user wants to export to Excel
# print("\n" + "="*50)
# export_choice = input("Export results to Excel? (y/n): ").lower().strip()
# if export_choice in ['y', 'yes']:
#     # Get boss name from boss_class
#     boss_name = boss_class.__name__
#     export_to_excel(results, boss_name, num_players, player_specs, num_sims)
# else:
#     print("Results not exported.")

